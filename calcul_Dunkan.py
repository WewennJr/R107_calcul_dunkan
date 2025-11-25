#### Importation des modules
import tkinter as tk
from tkinter import ttk
import matplotlib.pyplot as plt
import numpy as np
from PIL import Image, ImageTk
from openpyxl import load_workbook
import os
import sys

# --- CONFIGURATION DES POLICES ---
FONT_TEXTE = ("Arial", 14)
FONT_GRAS = ("Arial", 14, "bold")
FONT_TITRE = ("Arial", 22, "bold")
FONT_BOUTON = ("Arial", 16, "bold")

# --- Utilitaire pour le chemin du fichier ---
def get_excel_path():
    dossier_courant = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(dossier_courant, "Coef_Excel.xlsx")


# --- CHARGEMENT DES LABELS (R101, SAE...) ---
def charger_labels_depuis_excel():
    fichier = get_excel_path()
    liste_res = []
    liste_sae = []

    try:
        wb = load_workbook(fichier, data_only=True)
        ws = wb["Feuil1"]

        for row in ws.iter_rows(values_only=True):
            if row and len(row) >= 4 and isinstance(row[1], (int, float)):
                label = str(row[0])

                if label.upper().startswith("R"):
                    liste_res.append(label)
                elif label.upper().startswith("S"):
                    liste_sae.append(label)
                else:
                    liste_res.append(label)

        wb.close()
        return liste_res, liste_sae

    except FileNotFoundError:
        print(f"ERREUR : Le fichier {fichier} est introuvable.")
        return [], []
    except Exception as e:
        print(f"Erreur lecture labels : {e}")
        return [], []


# --- COMPTEUR DU NOMBRE DE COEFFICIENTS ---
def compter_coeffs_excel():
    try:
        wb = load_workbook(get_excel_path(), data_only=True)
        ws = wb["Feuil1"]

        nb = sum(
            1 for row in ws.iter_rows(values_only=True)
            if row and len(row) >= 4 and isinstance(row[1], (int, float))
        )

        wb.close()
        return nb
    except:
        return None


# --- Fonction de calcul des moyennes ---
def coef2(notes):
    nom_fichier_excel = get_excel_path()

    coefUE1, coefUE2, coefUE3 = [], [], []

    try:
        wb = load_workbook(nom_fichier_excel, data_only=True)
        ws = wb["Feuil1"]

        for row in ws.iter_rows(values_only=True):
            if row and len(row) >= 4 and isinstance(row[1], (int, float)):
                coefUE1.append(row[1])
                coefUE2.append(row[2])
                coefUE3.append(row[3])

        wb.close()
    except Exception as e:
        print(f"Erreur Excel : {e}")
        return None, None, None, None

    try:
        S1 = np.average(notes, weights=coefUE1)
        S2 = np.average(notes, weights=coefUE2)
        S3 = np.average(notes, weights=coefUE3)
    except Exception as e:
        print(f"Erreur calcul numpy : {e}")
        return None, None, None, None

    def get_color(value):
        if value >= 10: return "#4CAF50"
        elif value >= 8: return "#FF9800"
        else: return "#F44336"

    colors = [get_color(value) for value in [S1, S2, S3]]

    fig, ax = plt.subplots(figsize=(5, 4), facecolor="white")
    bars = ax.bar(["UE1", "UE2", "UE3"], [S1, S2, S3], width=0.5, color=colors)

    ax.set_title("Moyennes par UE", fontsize=14, fontweight='bold')
    ax.tick_params(axis='both', which='major', labelsize=12)
    ax.set_ylim(0, 20)

    for bar in bars:
        yval = bar.get_height()
        ax.text(bar.get_x() + bar.get_width() / 2, yval + 0.2, round(yval, 2),
                ha='center', va='bottom', fontsize=12, fontweight='bold')

    dossier_courant = os.path.dirname(os.path.abspath(__file__))
    chemin_graphique = os.path.join(dossier_courant, "Diagrame.jpg")

    plt.tight_layout()
    fig.savefig(chemin_graphique, dpi=100, bbox_inches="tight")
    plt.close(fig)

    return S1, S2, S3, chemin_graphique


# --- INITIALISATION DES DONNÉES ---
ressources, sae = charger_labels_depuis_excel()

if not ressources and not sae:
    ressources = ["Erreur"]
    sae = ["Fichier ?"]


# --- INTERFACE GRAPHIQUE ---
fenetre = tk.Tk()
fenetre.title("Calculateur de Moyennes R&T")
fenetre.configure(bg="#f0f4f8")
fenetre.geometry("1300x900")

fenetre.columnconfigure(0, weight=1)
fenetre.columnconfigure(1, weight=1)
fenetre.rowconfigure(1, weight=1)

tk.Label(
    fenetre,
    text="Calcul des Moyennes UE (Via Excel)",
    font=FONT_TITRE,
    bg="#f0f4f8",
    fg="#333"
).grid(row=0, column=0, columnspan=2, pady=30)

frame_gauche = tk.Frame(fenetre, bg="#f0f4f8")
frame_gauche.grid(row=1, column=0, sticky="nsew", padx=40)

entrees = []

tk.Label(frame_gauche, text="Ressources", font=FONT_BOUTON,
         bg="#f0f4f8", fg="#0056b3").grid(row=0, column=0, pady=15)

for i, res in enumerate(ressources):
    tk.Label(frame_gauche, text=res, font=FONT_GRAS, bg="#f0f4f8") \
        .grid(row=i + 1, column=0, padx=10, pady=4, sticky="e")
    entree = tk.Entry(frame_gauche, width=8, font=FONT_TEXTE, justify="center")
    entree.grid(row=i + 1, column=1, padx=10, pady=4)
    entrees.append(entree)

tk.Label(frame_gauche, text="SAE", font=FONT_BOUTON,
         bg="#f0f4f8", fg="#0056b3").grid(row=0, column=2, pady=15, padx=(50, 0))

for i, s in enumerate(sae):
    tk.Label(frame_gauche, text=s, font=FONT_GRAS, bg="#f0f4f8") \
        .grid(row=i + 1, column=2, padx=(50, 10), pady=4, sticky="e")
    entree = tk.Entry(frame_gauche, width=8, font=FONT_TEXTE, justify="center")
    entree.grid(row=i + 1, column=3, padx=10, pady=4)
    entrees.append(entree)


# --- CADRE DROIT ---
frame_droit = tk.Frame(fenetre, bg="white", bd=2, relief="groove")
frame_droit.grid(row=1, column=1, sticky="nsew", padx=40, pady=40)
frame_droit.columnconfigure(0, weight=1)

resultat_label = tk.Label(
    frame_droit,
    text="En attente de validation...",
    font=FONT_TEXTE,
    bg="white",
    justify="left"
)
resultat_label.grid(row=0, column=0, pady=30, sticky="w", padx=30)

canvas = tk.Canvas(frame_droit, bg="white", highlightthickness=0)
canvas.grid(row=1, column=0, sticky="nsew", padx=20, pady=10)


# --- LOGIQUE PRINCIPALE ---
def recuperer_infos():
    # --- Vérification des notes ---
    notes = []
    for e in entrees:
        raw = e.get().replace(',', '.').strip()

        if raw == "":
            raw = "0"

        try:
            val = float(raw)
        except ValueError:
            resultat_label.config(text=f"Erreur : '{raw}' n'est pas un nombre valide.", fg="red")
            return

        if val < 0:
            resultat_label.config(text=f"Erreur : une note ne peut pas être négative ({val}).", fg="red")
            return

        if val > 20:
            resultat_label.config(text=f"Erreur : une note ne peut pas dépasser 20 ({val}).", fg="red")
            return

        notes.append(val)

    # --- Vérification du nombre de coefficients ---
    nb_coefs = compter_coeffs_excel()
    if nb_coefs is None:
        resultat_label.config(text="Erreur : impossible de lire les coefficients Excel.", fg="red")
        return

    if len(notes) != nb_coefs:
        resultat_label.config(
            text=f"Erreur : {len(notes)} notes saisies, mais {nb_coefs} coefficients dans Excel.",
            fg="red"
        )
        return

    # --- Calcul final ---
    S1, S2, S3, chemin_img = coef2(notes)

    if S1 is None:
        resultat_label.config(text="Erreur : Problème lors du calcul. Vérifiez le fichier Excel.", fg="red")
        return

    resultat_label.config(
        text=f"CALCUL RÉUSSI :\n\n"
             f"- UE 1 : {S1:.2f} / 20\n"
             f"- UE 2 : {S2:.2f} / 20\n"
             f"- UE 3 : {S3:.2f} / 20",
        fg="#333"
    )

    # --- Affichage de l'image ---
    canvas.delete("all")
    try:
        img = Image.open(chemin_img)
        img = img.resize((600, 450), Image.Resampling.LANCZOS)
        photo = ImageTk.PhotoImage(img)
        canvas.image = photo
        canvas.create_image(300, 225, image=photo)

    except Exception as e:
        resultat_label.config(text=f"Erreur : impossible d'afficher le graphique ({e})", fg="red")


btn_valider = tk.Button(
    frame_droit,
    text="VALIDER LES NOTES",
    font=FONT_BOUTON,
    bg="#4CAF50",
    fg="white",
    activebackground="green",
    relief="flat",
    padx=20,
    pady=10,
    command=recuperer_infos
)
btn_valider.grid(row=2, column=0, pady=30)

fenetre.mainloop()
